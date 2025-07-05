import os
import cv2
import torch
import torch.nn as nn
import numpy as np
from PIL import Image
import torchvision
from torchvision import transforms
import glob
from transformers import BertTokenizer, BertModel
from paddleocr import PaddleOCR
import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from datetime import datetime
import time
from u2net import U2NET  # Import the U2NET model
from sklearn.metrics import precision_recall_fscore_support, confusion_matrix, average_precision_score
#nohup python recognition_base_resnet18_droupout_0.4.py > /dev/null 2>&1 &
# Set device
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
print(f"Using device: {device}")

# Define paths
test_dir = "CURE_dataset_test"
output_dir = "CURE_dataset_test_base_processed"
classification_model_path = "model_results_20250331_053757/best_model.pth"
u2net_model_path = "./saved_models/u2net_ft_best.pth"
results_excel = "pill_recognition_results_base_connected_component.xlsx"
os.makedirs(output_dir, exist_ok=True)

ocr = PaddleOCR(use_angle_cls=True, lang='en', use_gpu=False)
# Initialize BERT tokenizer and model
tokenizer = BertTokenizer.from_pretrained('bert-base-uncased')
bert_model = BertModel.from_pretrained('bert-base-uncased').to(device)

# Define transformations for image processing
transform = transforms.Compose([
    transforms.ToTensor(),
    transforms.Normalize(mean=[0.485, 0.456, 0.406], std=[0.229, 0.224, 0.225])
])

# Define CNN model
class CNNModel(nn.Module):
    def __init__(self, output_size=196):
        super(CNNModel, self).__init__()
        self.model = torchvision.models.resnet18(pretrained=True)
        self.model.fc = nn.Linear(self.model.fc.in_features, output_size)

    def forward(self, x):
        return self.model(x)

# Define combined model for RGB, contour, texture, and imprinted text
class CombinedModel(nn.Module):
    def __init__(self):
        super(CombinedModel, self).__init__()
        self.rgb_model = CNNModel(output_size=128)
        self.contour_model = CNNModel(output_size=128)
        self.texture_model = CNNModel(output_size=128)
        self.text_model = nn.Linear(768, 128)  # 768: BERT embedding size
        self.fc = nn.Linear(128 * 4, 196)  # Combine all features for classification
        self.dropout = nn.Dropout(0.3)

    def forward(self, rgb, contour, texture, text):
        rgb_features = self.rgb_model(rgb)
        contour_features = self.contour_model(contour)
        texture_features = self.texture_model(texture)
        text_features = self.text_model(text).squeeze(1)

        # Concatenate features
        combined_features = torch.cat((rgb_features, contour_features, texture_features, text_features), dim=1)
        combined_features = self.dropout(combined_features)
        output = self.fc(combined_features)
        return output

def extract_label_from_filename(filename):
    """
    Extract the first number before an underscore from the filename
    Example: "2_bottom_107.jpg" -> 2
    """
    match = re.match(r'^(\d+)_', filename)
    if match:
        return int(match.group(1))
    return None

# Load U2NET model for background removal
def load_u2net_model():
    u2net_model = U2NET(in_ch=3, out_ch=1)
    try:
        state_dict = torch.load(u2net_model_path, map_location=device)
        u2net_model.load_state_dict(state_dict)
        print(f"Successfully loaded U2NET model from {u2net_model_path}")
    except Exception as e:
        print(f"Error loading U2NET model: {str(e)}")
    
    u2net_model.to(device)
    u2net_model.eval()
    return u2net_model

# Function to remove background using U2NET
def remove_background_u2net(image, u2net_model):
    # Convert to RGB if needed
    if isinstance(image, np.ndarray):
        if image.shape[2] == 3:  # If it's already RGB
            rgb_image = image
        else:
            rgb_image = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)
        pil_image = Image.fromarray(rgb_image)
    else:
        pil_image = image
    
    # Resize to 320x320 as in TestDataset
    resized_img = pil_image.resize((320, 320))
    
    # Prepare image for U2NET
    img_tensor = transforms.Compose([
        transforms.ToTensor(),
        transforms.Normalize(mean=[0.485, 0.456, 0.406], std=[0.229, 0.224, 0.225])
    ])(resized_img).unsqueeze(0).to(device)
    
    # Get prediction
    with torch.no_grad():
        d0, _, _, _, _, _, _ = u2net_model(img_tensor)
    
    # Convert prediction to mask
    pred = d0.squeeze(0).squeeze(0).cpu().numpy()
    
    # Normalize to 0-1
    mask = (pred - pred.min()) / (pred.max() - pred.min())
    
    # Resize mask back to original size
    original_size = pil_image.size
    mask = cv2.resize(mask, original_size[::-1])
    
    # Create RGBA image with transparency
    rgba = np.zeros((original_size[1], original_size[0], 4), dtype=np.uint8)
    rgba[:, :, :3] = np.array(pil_image)
    rgba[:, :, 3] = (mask * 255).astype(np.uint8)
    
    return Image.fromarray(rgba, mode='RGBA')

# Function to load the trained models
def load_models():
    # Load classification model
    classification_model = CombinedModel()
    try:
        checkpoint = torch.load(classification_model_path, map_location=device)
    except Exception as e:
        print(f"Warning when loading classification model: {e}")
        checkpoint = torch.load(classification_model_path, map_location=device, weights_only=False)
    
    # Check if the checkpoint has 'model_state_dict' key
    if 'model_state_dict' in checkpoint:
        classification_model.load_state_dict(checkpoint['model_state_dict'])
    else:
        # Assume it's just the model state dict
        classification_model.load_state_dict(checkpoint)
    
    classification_model.to(device)
    classification_model.eval()
    
    # Load U2NET model for background removal
    u2net_model = load_u2net_model()
    
    return classification_model, u2net_model

# Function to convert text to tensor using BERT
def text_to_tensor(text):
    encoded_text = tokenizer(text, padding='max_length', truncation=True, max_length=128, return_tensors='pt')
    with torch.no_grad():
        bert_output = bert_model(**{key: val.to(device) for key, val in encoded_text.items()})
    return bert_output.last_hidden_state[:, 0, :]  # Use [CLS] token's embedding

# Function to extract image features
# Sửa lại chỉ phần trích xuất đặc trưng trong hàm extract_features
def extract_features(image_path):
    try:
        # Đọc và resize ảnh
        image = cv2.imread(image_path)
        if image is None:
            print(f"Không thể đọc ảnh: {image_path}")
            image = np.zeros((224, 224, 3), dtype=np.uint8)
            #return None, None, None, None
        
        # Giữ nguyên resize về 224x224 như trong huấn luyện
        image = cv2.resize(image, (224, 224))
        
        # 1. Trích xuất đặc trưng RGB - Sửa theo cách làm trong huấn luyện
        rgb_image = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)
        pil_image = Image.fromarray(rgb_image)
        rgb_tensor = transform(pil_image).unsqueeze(0).to(device)
        
        # 2. Trích xuất đặc trưng contour - Sửa theo cách làm trong huấn luyện
        gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
        blurred = cv2.GaussianBlur(gray, (5, 5), 0)
        edges = cv2.Canny(blurred, 50, 150)
        contours, _ = cv2.findContours(edges, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        contour_image = np.zeros_like(image)
        cv2.drawContours(contour_image, contours, -1, (255, 255, 255), 2)
        contour_pil = Image.fromarray(contour_image)
        contour_tensor = transform(contour_pil).unsqueeze(0).to(device)
        
        # 3. Trích xuất đặc trưng texture - Sửa theo cách làm trong huấn luyện
        texture_image = np.zeros_like(image)
        
        angles = [0, 45, 90, 135]
        for theta in angles:
            theta_rad = theta * np.pi / 180
            kernel = cv2.getGaborKernel((21, 21), 5.0, theta_rad, 10.0, 0.5, 0, ktype=cv2.CV_32F)
            filtered = cv2.filter2D(gray, cv2.CV_8UC3, kernel)
            texture_image[:,:,0] = np.maximum(texture_image[:,:,0], filtered)
            texture_image[:,:,1] = np.maximum(texture_image[:,:,1], filtered)
            texture_image[:,:,2] = np.maximum(texture_image[:,:,2], filtered)
        texture_pil = Image.fromarray(texture_image)
        texture_tensor = transform(texture_pil).unsqueeze(0).to(device)
        
        # 4. Trích xuất text thông qua OCR (giữ nguyên)
        try:
            ocr_result = ocr.ocr(image_path, cls=True)
            imprinted_text = ' '.join([line[1][0] for line in ocr_result[0]]) if ocr_result and ocr_result[0] else ""
            #print("imprint la", imprinted_text)  # Giữ nguyên debug print
        except Exception as e:
            print(f"Lỗi OCR trên {image_path}: {e}")
            imprinted_text = ""
        
        # Chuyển text thành tensor
        text_tensor = text_to_tensor(imprinted_text)
        
        return rgb_tensor, contour_tensor, texture_tensor, text_tensor
        
    except Exception as e:
        print(f"Lỗi khi xử lý {image_path}: {e}")
        default_tensor = torch.zeros((3, 224, 224), dtype=torch.float32)
        default_text = text_to_tensor("")
        return default_tensor, default_tensor, default_tensor, default_text
def process_pill_bounding_box(image, alpha_channel):
    """
    Xác định bounding box chính xác cho viên thuốc từ alpha channel
    Tối ưu cho viên thuốc nhỏ và không sử dụng padding
    """
    # Giảm ngưỡng để không bỏ sót viên thuốc nhỏ
    _, binary_mask = cv2.threshold(alpha_channel, 5, 255, cv2.THRESH_BINARY)
    
    # Áp dụng morphological operations nhẹ nhàng
    kernel = np.ones((3, 3), np.uint8)  # Kernel nhỏ hơn để không ảnh hưởng đến viên thuốc nhỏ
    binary_mask = cv2.morphologyEx(binary_mask, cv2.MORPH_CLOSE, kernel, iterations=1)
    
    # Tìm connected components - chỉ 1 connected component trên mỗi ảnh sau khi loại nền
    num_labels, labels, stats, centroids = cv2.connectedComponentsWithStats(binary_mask, connectivity=8)
    
    # Tìm connected component lớn nhất (bỏ qua background - index 0)
    largest_component_idx = 0
    largest_component_area = 0
    
    for i in range(1, num_labels):  # Bỏ qua background (index 0)
        area = stats[i, cv2.CC_STAT_AREA]
        if area > largest_component_area:
            largest_component_area = area
            largest_component_idx = i
    
    # Giảm kích thước tối thiểu xuống để không bỏ sót viên thuốc nhỏ
    if largest_component_idx > 0 and largest_component_area > 5:  # Giảm từ 100 xuống 20
        # Lấy bounding box từ stats
        x = stats[largest_component_idx, cv2.CC_STAT_LEFT]
        y = stats[largest_component_idx, cv2.CC_STAT_TOP]
        w = stats[largest_component_idx, cv2.CC_STAT_WIDTH]
        h = stats[largest_component_idx, cv2.CC_STAT_HEIGHT]
        
        # Không thêm padding
        return x, y, w, h
    
    # Fallback: sử dụng contour nếu connected components không hoạt động tốt
    contours, _ = cv2.findContours(binary_mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    
    if contours:
        # Sử dụng tất cả các contour thay vì chỉ lấy contour lớn nhất
        all_contours = np.vstack([cnt for cnt in contours])
        x, y, w, h = cv2.boundingRect(all_contours)
        
        return x, y, w, h
    
    # Fallback cuối cùng: nếu không tìm thấy contour, sử dụng phân tích pixel
    y_indices, x_indices = np.where(binary_mask > 0)
    
    if len(y_indices) > 0 and len(x_indices) > 0:
        x_min = np.min(x_indices)
        y_min = np.min(y_indices)
        x_max = np.max(x_indices)
        y_max = np.max(y_indices)
        
        w = x_max - x_min + 1
        h = y_max - y_min + 1
        
        return x_min, y_min, w, h
    else:
        # Nếu không tìm thấy pixel, trả về toàn bộ ảnh
        h_img, w_img = binary_mask.shape
        return 0, 0, w_img, h_img

def process_image(image_path, classification_model, u2net_model):
    try:
        # Extract true label from the filename
        filename = os.path.basename(image_path)
        true_label = extract_label_from_filename(filename)
        
        print(f"Processing {filename}, extracted label: {true_label}")
        
        # Read original image
        original_image = cv2.imread(image_path)
        if original_image is None:
            print(f"Failed to read image: {image_path}")
            return {
                'image': filename,
                'true_label': true_label,
                'prediction': None,
                'correct': False,
                'error': 'Failed to read image',
                'processing_time': 0
            }
        
        start_time = time.time()
        
        # Convert image for background removal
        pil_image = Image.fromarray(cv2.cvtColor(original_image, cv2.COLOR_BGR2RGB))
        
        # Kiểm tra nếu là file .bmp để thiết lập resolution là 96 dpi
        is_bmp = image_path.lower().endswith('.bmp')
        if is_bmp:
            dpi = (96, 96)
            pil_image.info['dpi'] = dpi
        
        # Remove background using U2NET
        no_bg_image = remove_background_u2net(pil_image, u2net_model)
        no_bg_array = np.array(no_bg_image)
        
        # Đảm bảo rằng ảnh đã loại nền có 4 kênh (RGBA)
        if no_bg_array.shape[2] == 4:
            # Lấy kênh alpha để xác định bounding box
            alpha_channel = no_bg_array[:, :, 3]
            
            # Tạo RGB với nền đen
            rgb_image = np.zeros((no_bg_array.shape[0], no_bg_array.shape[1], 3), dtype=np.uint8)
            alpha = no_bg_array[:, :, 3] / 255.0
            for c in range(3):
                rgb_image[:, :, c] = (no_bg_array[:, :, c] * alpha).astype(np.uint8)
        else:
            # Trường hợp hiếm gặp - ảnh không có kênh alpha
            rgb_image = no_bg_array[:, :, :3]
            gray = cv2.cvtColor(rgb_image, cv2.COLOR_RGB2GRAY)
            _, alpha_channel = cv2.threshold(gray, 10, 255, cv2.THRESH_BINARY)
        
        # Xác định bounding box từ alpha channel - không thêm padding
        x, y, w, h = process_pill_bounding_box(rgb_image, alpha_channel)
        
        # Kiểm tra nếu bounding box có kích thước hợp lệ
        if w <= 0 or h <= 0:
            processing_time = time.time() - start_time
            return {
                'image': filename,
                'true_label': true_label,
                'prediction': None,
                'correct': False,
                'error': f'Invalid bounding box: {[x, y, w, h]}',
                'processing_time': processing_time,
                'output_path': None
            }
        
        # Cắt ảnh RGB theo bounding box
        cropped_rgb = rgb_image[y:y+h, x:x+w]
        
        # Chuyển sang PIL Image để dễ xử lý
        pil_cropped = Image.fromarray(cropped_rgb)
        
        # Resize trực tiếp về 256x256 không giữ tỷ lệ
        pil_resized = pil_cropped.resize((256, 256), Image.LANCZOS)
        
        # Đảm bảo thông tin dpi cho file .bmp
        if is_bmp:
            pil_resized.info['dpi'] = (96, 96)
        
        # Lưu ảnh cuối cùng
        output_path = os.path.join(output_dir, os.path.basename(image_path).replace('.jpg', '.png').replace('.bmp', '.png'))
        pil_resized.save(output_path, dpi=(96, 96) if is_bmp else None)
        
        # Trích xuất đặc trưng từ ảnh đã xử lý
        rgb_tensor, contour_tensor, texture_tensor, text_tensor = extract_features(output_path)
        
        if rgb_tensor is not None:
            # Đưa ra dự đoán
            with torch.no_grad():
                outputs = classification_model(rgb_tensor, contour_tensor, texture_tensor, text_tensor)
                _, predicted = torch.max(outputs, 1)
                prediction = predicted.item()
            
            processing_time = time.time() - start_time
            is_correct = prediction == true_label
            
            print(f"Predicted: {prediction}, True label: {true_label}, Correct: {is_correct}, Time: {processing_time:.2f}s")
            
            # Trả về kết quả
            return {
                'image': filename,
                'true_label': true_label,
                'prediction': prediction,
                'correct': is_correct,
                'error': None,
                'processing_time': processing_time,
                'output_path': output_path
            }
        else:
            processing_time = time.time() - start_time
            return {
                'image': filename,
                'true_label': true_label,
                'prediction': None,
                'correct': False,
                'error': 'Feature extraction failed',
                'processing_time': processing_time,
                'output_path': output_path if 'output_path' in locals() else None
            }
    
    except Exception as e:
        processing_time = time.time() - start_time if 'start_time' in locals() else 0
        return {
            'image': os.path.basename(image_path),
            'true_label': true_label if 'true_label' in locals() else None,
            'prediction': None,
            'correct': False,
            'error': str(e),
            'processing_time': processing_time,
            'output_path': None
        }
def save_to_excel(results, excel_path):
    # Create DataFrame
    data = []
    for result in results:
        data.append({
            'Image': result['image'],
            'True Label': result['true_label'],
            'Predicted Label': result['prediction'],
            'Correct': 'Yes' if result['correct'] else 'No',
            'Error': result['error'] if result['error'] else '',
            'Processing Time (s)': round(result['processing_time'], 2),
            'Output Path': result['output_path'] if result['output_path'] else ''
        })
    
    df = pd.DataFrame(data)
    
    # Calculate summary statistics
    total_images = len(df)
    successful_predictions = df['Predicted Label'].notnull().sum()
    correct_predictions = (df['Correct'] == 'Yes').sum()
    accuracy = correct_predictions / total_images if total_images > 0 else 0
    avg_processing_time = df['Processing Time (s)'].mean()
    error_count = df['Error'].notnull().sum()
    
    # Calculate precision, recall, f1 score
    from sklearn.metrics import precision_recall_fscore_support, confusion_matrix, average_precision_score
    
    # Filter out entries without predictions
    valid_df = df.dropna(subset=['Predicted Label'])
    
    # Extract all unique classes from true and predicted labels
    # First method: Get classes from the data
    data_classes = sorted(set(valid_df['True Label'].dropna().tolist() + 
                           valid_df['Predicted Label'].dropna().tolist()))
    
    # Second method: Ensure we include all possible classes from 0 to 195
    num_classes = 196  # Total number of classes the model was trained on
    all_possible_classes = list(range(num_classes))
    
    # Use all possible classes to ensure we analyze all 196 classes
    unique_classes = all_possible_classes
    
    if len(valid_df) > 0:
        y_true = valid_df['True Label'].values
        y_pred = valid_df['Predicted Label'].values
        
        precision, recall, f1, _ = precision_recall_fscore_support(
            y_true, y_pred, average='weighted', zero_division=0
        )
        
        # Calculate mAP
        try:
            # Convert to one-hot format for mAP calculation
            num_classes = 196  # Assuming 196 classes as in the training code
            one_hot_labels = np.zeros((len(y_true), num_classes))
            for i, label in enumerate(y_true):
                if 0 <= label < num_classes:
                    one_hot_labels[i, int(label)] = 1
            
            # Get model prediction scores
            # Since we don't have actual scores, we'll create dummy scores (1.0 for predicted class)
            pred_scores = np.zeros((len(y_pred), num_classes))
            for i, pred in enumerate(y_pred):
                if 0 <= pred < num_classes:
                    pred_scores[i, int(pred)] = 1.0
            
            # Calculate per-class mAP and weighted average
            mAP_per_class = []
            class_support = []
            
            for i in range(num_classes):
                if np.sum(one_hot_labels[:, i]) > 0:
                    ap = average_precision_score(one_hot_labels[:, i], pred_scores[:, i])
                    support = np.sum(one_hot_labels[:, i])
                    mAP_per_class.append(ap)
                    class_support.append(support)
            
            mAP = np.average(mAP_per_class, weights=class_support) if mAP_per_class else 0.0
        except Exception as e:
            print(f"Error in mAP calculation: {e}")
            mAP = 0.0
    else:
        precision, recall, f1, mAP = 0.0, 0.0, 0.0, 0.0
    
    # Create summary DataFrame
    summary_data = {
        'Metric': ['Total Images', 'Successfully Processed', 'Correct Predictions', 'Accuracy', 
                  'Precision', 'Recall', 'F1 Score', 'mAP',
                  'Average Processing Time (s)', 'Errors'],
        'Value': [total_images, successful_predictions, correct_predictions, 
                 f"{accuracy:.2%}", f"{precision:.4f}", f"{recall:.4f}", f"{f1:.4f}", f"{mAP:.4f}",
                 f"{avg_processing_time:.2f}", error_count]
    }
    summary_df = pd.DataFrame(summary_data)
    
    # Save to Excel with multiple sheets
    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        # Sheet 1: Summary
        summary_df.to_excel(writer, sheet_name='Summary', index=False)
        
        # Sheet 2: Detailed Results
        df.to_excel(writer, sheet_name='Detailed Results', index=False)
        
        # Sheet 3: Accuracy by Class Analysis
        if len(valid_df) > 0 and len(unique_classes) > 0:
            # Create DataFrame for class-based accuracy analysis
            accuracy_by_class = []
            for class_label in unique_classes:
                # Filter samples for this class
                class_samples = valid_df[valid_df['True Label'] == class_label]
                class_total = len(class_samples)
                class_correct = sum(class_samples['Predicted Label'] == class_label)
                class_accuracy = class_correct / class_total if class_total > 0 else 0
                
                # Only include classes that have samples or predictions in the report
                if class_total > 0 or sum(valid_df['Predicted Label'] == class_label) > 0:
                
                    accuracy_by_class.append({
                        'Class': class_label,
                        'Total Samples': class_total,
                        'Correct Predictions': class_correct,
                        'Incorrect Predictions': class_total - class_correct,
                        'Accuracy': class_accuracy
                    })
            
            accuracy_df = pd.DataFrame(accuracy_by_class)
            # Sort by accuracy descending
            accuracy_df = accuracy_df.sort_values(by='Accuracy', ascending=False)
            accuracy_df.to_excel(writer, sheet_name='Accuracy by Class', index=False)
        
        # Sheet 4: Precision by Class Analysis
        if len(valid_df) > 0 and len(unique_classes) > 0:
            # Calculate Precision for each class
            precision_by_class = []
            for class_label in unique_classes:
                # Samples predicted as this class
                predicted_as_class = valid_df[valid_df['Predicted Label'] == class_label]
                total_predicted = len(predicted_as_class)
                # Samples correctly predicted as this class
                true_positives = sum(predicted_as_class['True Label'] == class_label)
                
                class_precision = true_positives / total_predicted if total_predicted > 0 else 0
                
                # Include this class in the report if it has samples in the test set
                # or if any samples were predicted as this class
                if total_predicted > 0 or sum(valid_df['True Label'] == class_label) > 0:
                    precision_by_class.append({
                        'Class': class_label,
                        'True Positives': true_positives,
                        'Total Predicted as This Class': total_predicted,
                        'False Positives': total_predicted - true_positives,
                        'Precision': class_precision,
                        'Formula': 'TP / (TP + FP)'
                    })
            
            precision_df = pd.DataFrame(precision_by_class)
            precision_df = precision_df.sort_values(by='Precision', ascending=False)
            precision_df.to_excel(writer, sheet_name='Precision by Class', index=False)
        
        # Sheet 5: Recall by Class Analysis
        if len(valid_df) > 0 and len(unique_classes) > 0:
            # Calculate Recall for each class
            recall_by_class = []
            for class_label in unique_classes:
                # Samples actually in this class
                actual_class_samples = valid_df[valid_df['True Label'] == class_label]
                total_actual = len(actual_class_samples)
                # Samples correctly predicted as this class
                true_positives = sum(actual_class_samples['Predicted Label'] == class_label)
                
                class_recall = true_positives / total_actual if total_actual > 0 else 0
                
                # Include this class in the report if it has samples in the test set
                # or if any samples were predicted as this class
                if total_actual > 0 or sum(valid_df['Predicted Label'] == class_label) > 0:
                    recall_by_class.append({
                        'Class': class_label,
                        'True Positives': true_positives,
                        'Total Actual in This Class': total_actual,
                        'False Negatives': total_actual - true_positives,
                        'Recall': class_recall,
                        'Formula': 'TP / (TP + FN)'
                    })
            
            recall_df = pd.DataFrame(recall_by_class)
            recall_df = recall_df.sort_values(by='Recall', ascending=False)
            recall_df.to_excel(writer, sheet_name='Recall by Class', index=False)
        
        # Sheet 6: F1 Score by Class Analysis
        if len(valid_df) > 0 and len(unique_classes) > 0:
            # Calculate F1 Score for each class
            f1_by_class = []
            for class_label in unique_classes:
                # Get precision and recall from previously calculated DataFrames
                precision_item = next((item for item in precision_by_class if item['Class'] == class_label), None)
                recall_item = next((item for item in recall_by_class if item['Class'] == class_label), None)
                
                # Only include classes that have entries in either precision or recall lists
                if precision_item is not None or recall_item is not None:
                    class_precision = precision_item['Precision'] if precision_item is not None else 0
                    class_recall = recall_item['Recall'] if recall_item is not None else 0
                    
                    # Calculate F1 Score
                    if class_precision + class_recall > 0:
                        class_f1 = 2 * (class_precision * class_recall) / (class_precision + class_recall)
                    else:
                        class_f1 = 0
                    
                    f1_by_class.append({
                        'Class': class_label,
                        'Precision': class_precision,
                        'Recall': class_recall,
                        'F1 Score': class_f1,
                        'Formula': '2 * (Precision * Recall) / (Precision + Recall)'
                    })
            
            f1_df = pd.DataFrame(f1_by_class)
            f1_df = f1_df.sort_values(by='F1 Score', ascending=False)
            f1_df.to_excel(writer, sheet_name='F1 Score by Class', index=False)
        
        # Sheet 7: mAP by Class Analysis
        if len(valid_df) > 0 and len(unique_classes) > 0:
            # Calculate Average Precision for each class
            ap_by_class = []
            for class_label in unique_classes:
                if isinstance(class_label, (int, float)) and 0 <= class_label < num_classes:
                    # Get the counts for this class
                    class_support = sum(y_true == class_label)
                    tp = sum((y_true == class_label) & (y_pred == class_label))
                    fp = sum((y_true != class_label) & (y_pred == class_label))
                    fn = sum((y_true == class_label) & (y_pred != class_label))
                    
                    # Only calculate AP if this class has samples in the dataset or was predicted
                    if class_support > 0 or sum(y_pred == class_label) > 0:
                        # Create one-hot label for this class
                        true_label = np.zeros(len(y_true))
                        true_label[y_true == class_label] = 1
                        
                        # Create prediction scores
                        pred_score = np.zeros(len(y_pred))
                        pred_score[y_pred == class_label] = 1
                        
                        # Calculate Average Precision
                        try:
                            ap = average_precision_score(true_label, pred_score)
                        except Exception as e:
                            print(f"Error calculating AP for class {class_label}: {e}")
                            ap = 0
                        
                        ap_by_class.append({
                            'Class': class_label,
                            'Average Precision': ap,
                            'Support': class_support,
                            'True Positives': tp,
                            'False Positives': fp,
                            'False Negatives': fn
                        })
            
            if ap_by_class:  # Only create the sheet if we have data
                ap_df = pd.DataFrame(ap_by_class)
                ap_df = ap_df.sort_values(by='Average Precision', ascending=False)
                ap_df.to_excel(writer, sheet_name='mAP by Class', index=False)
        
        # Sheet 8: Confusion Matrix
        if len(valid_df) > 1:
            try:
                # Get unique classes that actually appear in the data
                unique_classes_in_data = sorted(set(y_true) | set(y_pred))
                
                if len(unique_classes_in_data) > 1:
                    # Calculate confusion matrix only for classes in the data
                    cm = confusion_matrix(y_true, y_pred, labels=unique_classes_in_data)
                    cm_df = pd.DataFrame(cm, index=unique_classes_in_data, columns=unique_classes_in_data)
                    cm_df.index.name = 'True Label'
                    cm_df.columns.name = 'Predicted Label'
                    cm_df.to_excel(writer, sheet_name='Confusion Matrix')
            except Exception as e:
                print(f"Error creating confusion matrix: {e}")
        
        # Sheet 9: Metrics Explanation
        metrics_explanation = pd.DataFrame({
            'Metric': ['Accuracy', 'Precision', 'Recall', 'F1 Score', 'mAP'],
            'Formula': [
                'Correct Predictions / Total Images',
                'TP / (TP + FP)',
                'TP / (TP + FN)',
                '2 * (Precision * Recall) / (Precision + Recall)',
                'Area under the Precision-Recall Curve'
            ],
            'Explanation': [
                'Proportion of correctly classified samples across all samples',
                'Proportion of samples predicted as class X that actually belong to class X',
                'Proportion of samples of class X that are correctly predicted as class X',
                'Harmonic mean of Precision and Recall',
                'Weighted average of Average Precision across all classes'
            ]
        })
        metrics_explanation.to_excel(writer, sheet_name='Metrics Explanation', index=False)
        
        # Format sheets
        workbook = writer.book
        
        # Format summary sheet
        summary_sheet = writer.sheets['Summary']
        for i in range(1, len(summary_data['Metric']) + 1):
            summary_sheet.cell(row=i+1, column=1).font = Font(bold=True)
        
        # Format detailed results sheet
        results_sheet = writer.sheets['Detailed Results']
        for i in range(1, len(df.columns) + 1):
            results_sheet.cell(row=1, column=i).font = Font(bold=True)
        
        # Highlight correct/incorrect predictions
        for i, correct in enumerate(df['Correct'], start=2):
            if correct == 'Yes':
                results_sheet.cell(row=i, column=4).fill = PatternFill(start_color="C6EFCE", fill_type="solid")
            else:
                results_sheet.cell(row=i, column=4).fill = PatternFill(start_color="FFC7CE", fill_type="solid")
# Main function to process all test images
# Function to process all test images
def process_test_images():
    # Load models
    classification_model, u2net_model = load_models()
    
    # Find all images in the test directory
    image_files = []
    for ext in ['*.jpg', '*.bmp']:
        image_files.extend(glob.glob(os.path.join(test_dir, ext)))
    
    print(f"Found {len(image_files)} images to process")
    
    results = []
    
    # Process each image
    for image_file in image_files:
        result = process_image(image_file, classification_model, u2net_model)
        results.append(result)
    
    # Calculate overall accuracy
    correct_predictions = sum(1 for r in results if r['correct'])
    total_with_predictions = sum(1 for r in results if r['prediction'] is not None)
    total_images = len(results)
    
    accuracy = correct_predictions / total_with_predictions if total_with_predictions > 0 else 0
    
    # Extract all true labels and predictions for metric calculation
    true_labels = [r['true_label'] for r in results if r['prediction'] is not None]
    predictions = [r['prediction'] for r in results if r['prediction'] is not None]
    
    # Calculate metrics using sklearn
    from sklearn.metrics import precision_recall_fscore_support, average_precision_score
    
    if true_labels and predictions:
        precision, recall, f1, _ = precision_recall_fscore_support(
            true_labels, predictions, average='weighted', zero_division=0
        )
        
        # Calculate mAP (simplified version)
        try:
            num_classes = 196
            one_hot_true = np.zeros((len(true_labels), num_classes))
            for i, label in enumerate(true_labels):
                if 0 <= label < num_classes:
                    one_hot_true[i, label] = 1
                    
            # Create simplified prediction scores (1.0 for predicted class)
            pred_scores = np.zeros((len(predictions), num_classes))
            for i, pred in enumerate(predictions):
                if 0 <= pred < num_classes:
                    pred_scores[i, pred] = 1.0
                    
            # Calculate mAP per class
            mAP_values = []
            for c in range(num_classes):
                if np.sum(one_hot_true[:, c]) > 0:  # Only for classes that appear in test set
                    class_ap = average_precision_score(one_hot_true[:, c], pred_scores[:, c])
                    mAP_values.append(class_ap)
                    
            mAP = np.mean(mAP_values) if mAP_values else 0.0
        except Exception as e:
            print(f"Error calculating mAP: {e}")
            mAP = 0.0
    else:
        precision, recall, f1, mAP = 0.0, 0.0, 0.0, 0.0
    
    # Print summary
    print(f"\nResults Summary:")
    print(f"Total images: {total_images}")
    print(f"Images with predictions: {total_with_predictions}")
    print(f"Correct predictions: {correct_predictions}")
    print(f"Accuracy: {accuracy:.2%}")
    print(f"Precision: {precision:.4f}")
    print(f"Recall: {recall:.4f}")
    print(f"F1 Score: {f1:.4f}")
    print(f"mAP: {mAP:.4f}")
    
    # Add metric results to each result item
    for r in results:
        r['overall_metrics'] = {
            'accuracy': accuracy,
            'precision': precision,
            'recall': recall,
            'f1': f1,
            'mAP': mAP
        }
    
    # Save results to Excel
    save_to_excel(results, results_excel)
    
    # Save a simple text summary
    with open(os.path.join(output_dir, 'results_summary.txt'), 'w') as f:
        f.write(f"Results Summary:\n")
        f.write(f"Total images: {total_images}\n")
        f.write(f"Images with predictions: {total_with_predictions}\n")
        f.write(f"Correct predictions: {correct_predictions}\n")
        f.write(f"Accuracy: {accuracy:.2%}\n")
        f.write(f"Precision: {precision:.4f}\n")
        f.write(f"Recall: {recall:.4f}\n")
        f.write(f"F1 Score: {f1:.4f}\n")
        f.write(f"mAP: {mAP:.4f}\n")

# Main function
if __name__ == "__main__":
    # Record start time for overall process
    overall_start_time = time.time()
    
    # Process all test images
    process_test_images()
    
    # Calculate and print total runtime
    total_runtime = time.time() - overall_start_time
    print(f"\nTotal runtime: {total_runtime:.2f} seconds ({total_runtime/60:.2f} minutes)")

